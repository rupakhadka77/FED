# coding=utf-8

import openpyxl
import numpy as np
from cleanText import cleanString
from sklearn.svm import LinearSVC
from sklearn.metrics import confusion_matrix, f1_score, precision_score, recall_score
from sklearn.model_selection import train_test_split
from sklearn.feature_extraction.text import CountVectorizer, TfidfVectorizer

# Get the original dataset
def store():

    workBookOld = openpyxl.load_workbook('DataSet.xlsx')
    dataSheetOld = workBookOld['Data set']

    xData = []
    yData = []

    rows = dataSheetOld.max_row

    for i in range(2, rows+1):

        if (str(dataSheetOld.cell(row = i, column = 2).value) != 'None'):
            xData.append(str(cleanString(dataSheetOld.cell(row = i, column = 1).value)))
            if (str(dataSheetOld.cell(row = i, column = 2).value) == "1"):
                yData.append(1)
            else:
                yData.append(0)

    # NOTE: to train data on the entire dataset, simply return xData and yData
    # Splitting the data like this is to obtain test cases and calculate the F-score of the learning algorithm
    xTrain, xTest, yTrain, yTest = train_test_split(xData, yData, test_size=0.2, random_state=0)
    return xTrain, xTest, yTrain, yTest


# Calculating the F-score
def calcFScore(xTest, yTest, model, vectorizer):
    #d_data= ['I am Kellim Worthington i am doing masters in public administration from the University of London i have done my graduation in business administration from one of the best university of United Kingdom. Well, i am here to ask you one thing, That i want to buy a laptop for my assignments because one day i was making my assignment and that time my laptop was on the charging and due to shortcircuit my laptop expired. Now I am so worried about my <a href="https://www.assignmentmaster.co.uk/" title="" target="_blank">Assignment Writing UK</a> so please let me know which laptop is the best in these days.  please suggest me because i want to purchase tomorrow.  ']
    #dataa=[1]
    xTestMatrix = vectorizer.transform(xTest)
    yTestMatrix = np.asarray(yTest)
    # print(xTestMatrix)
    # print(xTest[4])
    
    

    result = model.predict(xTestMatrix)
    matrix = confusion_matrix(yTestMatrix, result)

    fScore = f1_score(yTestMatrix, result, pos_label = 0)
    precision = precision_score(yTestMatrix, result, pos_label=0)
    recall = recall_score(yTestMatrix, result, pos_label=0)
    return fScore, precision, recall, matrix

# Test new data for Spam
def predict(emailBody, model, vectorizer):

    featureMatrix = vectorizer.transform([cleanString(emailBody)])
    result = model.predict(featureMatrix)
    print("Predicting...")

    if result > 0.8:
        return "Spam"
    else:
        return "Not Spam"

model = LinearSVC(class_weight='balanced')

# Create training data
xTrain, xTest, yTrain, yTest = store()

vectorizer = TfidfVectorizer(stop_words='english', max_df=75)
yTrainMatrix = np.asarray(yTrain)
xTrainMatrix = vectorizer.fit_transform(xTrain)

# Training SVM classifier
model.fit(xTrainMatrix, yTrainMatrix)
fScore, precision, recall, matrix = calcFScore(xTest, yTest, model, vectorizer)

print(fScore, precision, recall, matrix)

# spamData = "This code is genuinely helpful and executing the program.  Conversational, guidance and active forum are always the best directions to get a solution."
spamData = "WINNER!! As a valued network customer you have been selected to receivea å£900 prize reward! To claim call 09061701461. Claim code KL341. Valid 12 hours only."
label = predict(spamData, model, vectorizer)
print("Email is: %s" % label)